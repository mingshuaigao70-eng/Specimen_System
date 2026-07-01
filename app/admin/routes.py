from flask import Blueprint, render_template, request, redirect, url_for, flash, abort, current_app, session, jsonify
from flask_login import login_required, current_user
from functools import wraps
from datetime import datetime
from ..extensions import db
from ..models import User, Specimen, SpecimenImage, SpecimenCategory, PageContent
from ..utils.password import generate_scrypt_hash
from app.utils.file_util import FileHandler
from app.utils.excel_util import parse_workbook, extract_code_from_number
import json
import os
import re
import uuid
import zipfile
import shutil
from io import BytesIO
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import selectinload
from app.utils.time_utils import now , CHINA_TZ
from ..utils.search_utils import build_specimen_search_filter

admin_bp = Blueprint('admin', __name__, url_prefix='/admin')

# -------------------- 鏉冮檺瑁呴グ鍣?-------------------- #
def superadmin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if current_user.role != 'superadmin':
            abort(403)
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if current_user.role not in ['superadmin', 'admin']:
            abort(403)
        return f(*args, **kwargs)
    return decorated_function

# -------------------- 瀹夊叏琛ㄥ崟瑙ｆ瀽宸ュ叿 -------------------- #
def _parse_datetime_form_value(value_str, default=None):
    """瀹夊叏瑙ｆ瀽琛ㄥ崟涓殑鏃ユ湡鏃堕棿鍊硷紝鏍煎紡閿欒鏃惰繑鍥?default 鑰岄潪宕╂簝"""
    if not value_str:
        return default
    try:
        fmt = "%Y-%m-%dT%H:%M" if 'T' in value_str else "%Y-%m-%d"
        dt = datetime.strptime(value_str, fmt)
        return CHINA_TZ.localize(dt)
    except (ValueError, TypeError):
        return default


def _parse_float_form_value(value_str, default=None):
    """瀹夊叏瑙ｆ瀽琛ㄥ崟涓殑娴偣鏁板€硷紝鏍煎紡閿欒鏃惰繑鍥?default 鑰岄潪宕╂簝"""
    if not value_str or not str(value_str).strip():
        return default
    try:
        return float(value_str)
    except (ValueError, TypeError):
        return default


def _parse_int_form_value(value_str, default=None):
    """瀹夊叏瑙ｆ瀽琛ㄥ崟涓殑鏁存暟鍊硷紝鏍煎紡閿欒鏃惰繑鍥?default 鑰岄潪宕╂簝"""
    if not value_str or not str(value_str).strip():
        return default
    try:
        return int(value_str)
    except (ValueError, TypeError):
        return default


# -------------------- 骞冲彴棣栭〉 -------------------- #
@admin_bp.route('/dashboard')
@login_required
@admin_required
def dashboard():
    default_url = '/users' if current_user.role == 'superadmin' else '/specimens'
    return render_template('admin/admin_dashboard.html', default_url=default_url)

# -------------------- 椤甸潰閰嶇疆 -------------------- #
@admin_bp.route('/page_config', methods=['GET', 'POST'])
@login_required
@superadmin_required
def page_config():
    if request.method == 'POST':
        rows = PageContent.query.all()
        row_map = {f'{r.page}.{r.section}': r for r in rows}
        for key in request.form:
            if key in row_map:
                row_map[key].content = request.form[key]
                row_map[key].updated_by = current_user.username
                row_map[key].updated_at = datetime.now()

        # 澶勭悊棣栭〉妯箙鑳屾櫙鍥句笂浼?
        image_file = request.files.get('banner_image')
        if image_file and FileHandler.check_file(image_file):
            try:
                ext = image_file.filename.rsplit('.', 1)[1].lower()
                filename = f"banner.{ext}"
                # 鍒犻櫎鏃фí骞呭浘鐗?
                old = PageContent.query.filter_by(page='landing', section='banner_image').first()
                if old and old.content:
                    old_path = os.path.join(current_app.config['UPLOAD_FOLDER_CATEGORY_IMAGES'], os.path.basename(old.content))
                    if os.path.exists(old_path):
                        os.remove(old_path)
                relative_path = FileHandler.save_file(image_file, folder_key='UPLOAD_FOLDER_CATEGORY_IMAGES', filename=filename)
                if old:
                    old.content = relative_path
                    old.updated_by = current_user.username
                    old.updated_at = datetime.now()
                else:
                    db.session.add(PageContent(
                        page='landing', section='banner_image',
                        content=relative_path,
                        updated_by=current_user.username
                    ))
            except IOError as e:
                flash(f'妯箙鍥剧墖淇濆瓨澶辫触: {e}', 'error')
                return redirect(url_for('admin.page_config'))

        # 澶勭悊鍏充簬鎴戜滑椤甸潰鍦板浘鍥剧墖涓婁紶
        map_file = request.files.get('map_image')
        if map_file and FileHandler.check_file(map_file):
            try:
                ext = map_file.filename.rsplit('.', 1)[1].lower()
                filename = f"map.{ext}"
                old = PageContent.query.filter_by(page='about', section='map_image').first()
                if old and old.content:
                    old_path = os.path.join(current_app.config['UPLOAD_FOLDER_CATEGORY_IMAGES'], os.path.basename(old.content))
                    if os.path.exists(old_path):
                        os.remove(old_path)
                relative_path = FileHandler.save_file(map_file, folder_key='UPLOAD_FOLDER_CATEGORY_IMAGES', filename=filename)
                if old:
                    old.content = relative_path
                    old.updated_by = current_user.username
                    old.updated_at = datetime.now()
                else:
                    db.session.add(PageContent(
                        page='about', section='map_image',
                        content=relative_path,
                        updated_by=current_user.username
                    ))
            except IOError as e:
                flash(f'鍦板浘鍥剧墖淇濆瓨澶辫触: {e}', 'error')
                return redirect(url_for('admin.page_config'))

        db.session.commit()
        current_app.logger.info('page_config_updated by=%s', current_user.username)
        flash('页面内容已更新', 'success')
        return redirect(url_for('admin.page_config'))

    rows = PageContent.query.all()
    content = {f'{r.page}.{r.section}': r.content for r in rows}
    return render_template('admin/admin_page_config.html', content=content)

# ==================== 鐢ㄦ埛绠＄悊 ==================== #

ALLOWED_ROLES = ('admin', 'user')  # superadmin 涓嶅厑璁搁€氳繃 Web 鐣岄潰鍒嗛厤

# 甯歌寮卞瘑鐮侀粦鍚嶅崟 鈥?鍧囨弧瓒冲瘑鐮佸己搴﹁鍒欎絾鏋佹槗琚寽娴?
COMMON_WEAK_PASSWORDS = {
    'Password1!', 'Admin123!', 'Admin123!@#', 'admin123!@#',
    'Qwerty1!', 'Qwerty123!', 'Welcome1!', 'Welcome123!',
    'Changeme1!', 'P@ssw0rd1', 'Pa$$w0rd1', 'Abcd1234!',
    'ABCabc123!', 'Passw0rd!', 'Passw0rd1!',
}


def _validate_password(password):
    """鏍￠獙瀵嗙爜寮哄害锛岃繑鍥?(is_valid, error_message)"""
    if len(password) < 8:
        return False, '瀵嗙爜鑷冲皯 8 浣嶏紒'
    if not (any(c.islower() for c in password) and
            any(c.isupper() for c in password) and
            any(c.isdigit() for c in password) and
            any(not c.isalnum() for c in password)):
        return False, '瀵嗙爜闇€鍖呭惈澶у啓瀛楁瘝銆佸皬鍐欏瓧姣嶃€佹暟瀛椼€佺壒娈婂瓧绗︼紒'
    if password in COMMON_WEAK_PASSWORDS:
        return False, '密码过于常见，请使用更复杂的密码'
    return True, None


@admin_bp.route('/users')
@login_required
@superadmin_required
def manage_users():
    users = User.query.all()
    return render_template('admin/admin_user_management.html', users=users)

@admin_bp.route('/users/add', methods=['GET', 'POST'])
@login_required
@superadmin_required
def add_user():
    if request.method == 'POST':
        username = request.form['username'].strip()
        role = request.form['role']
        password = request.form['password']

        if role not in ALLOWED_ROLES:
            flash('无效的角色值', 'error')
            return redirect(url_for('admin.add_user'))

        if User.query.filter_by(username=username).first():
            flash('用户名已存在', 'error')
            return redirect(url_for('admin.add_user'))

        # 瀵嗙爜寮哄害鏍￠獙
        valid, err = _validate_password(password)
        if not valid:
            flash(err, 'error')
            return redirect(url_for('admin.add_user'))

        new_user = User(
            username=username,
            role=role,
            password_hash=generate_scrypt_hash(password)
        )
        db.session.add(new_user)
        db.session.commit()
        current_app.logger.info('user_created by=%s username=%s role=%s', current_user.username, username, role)
        flash('新增用户成功', 'success')
        return redirect(url_for('admin.manage_users'))
    return render_template('admin/admin_add_user.html')


@admin_bp.route('/users/edit/<int:user_id>', methods=['GET', 'POST'])
@login_required
@superadmin_required
def edit_user(user_id):
    user = User.query.get_or_404(user_id)
    if request.method == 'POST':
        # 鈹€鈹€ 瓒呯骇绠＄悊鍛樿处鍙蜂粎鏀寔淇敼瀵嗙爜 鈹€鈹€
        if user.role == 'superadmin':
            password = request.form.get('password', '').strip()
            if not password:
                flash('璇疯緭鍏ユ柊瀵嗙爜', 'error')
                return redirect(url_for('admin.manage_users'))
            valid, err = _validate_password(password)
            if not valid:
                flash(err, 'error')
                return redirect(url_for('admin.manage_users'))
            user.password_hash = generate_scrypt_hash(password)
            db.session.commit()
            current_app.logger.info('superadmin_password_updated by=%s target=%s', current_user.username, user.username)
            flash('瀵嗙爜宸叉洿鏂帮紒', 'success')
            return redirect(url_for('admin.manage_users'))

        # 鈹€鈹€ 闈炶秴绾х鐞嗗憳锛氬彲淇敼鐢ㄦ埛鍚嶃€佽鑹诧紝瀵嗙爜鍙€?鈹€鈹€
        new_username = request.form.get('username', '').strip()
        new_role = request.form.get('role', 'user')
        if new_role not in ALLOWED_ROLES:
            flash('无效的角色值', 'error')
            return redirect(url_for('admin.manage_users'))

        existing = User.query.filter(User.username == new_username, User.id != user_id).first()
        if existing:
            flash('用户名已存在', 'error')
            return redirect(url_for('admin.manage_users'))

        user.username = new_username
        user.role = new_role

        password = request.form.get('password', '').strip()
        if password:
            valid, err = _validate_password(password)
            if not valid:
                flash(err, 'error')
                return redirect(url_for('admin.manage_users'))
            user.password_hash = generate_scrypt_hash(password)

        db.session.commit()
        current_app.logger.info('user_updated by=%s target=%s role=%s', current_user.username, user.username, user.role)
        flash('鐢ㄦ埛淇℃伅宸叉洿鏂帮紒', 'success')
        return redirect(url_for('admin.manage_users'))
    return render_template('admin/admin_edit_user.html', user=user)

@admin_bp.route('/users/delete/<int:user_id>', methods=['POST'])
@login_required
@superadmin_required
def delete_user(user_id):
    user = User.query.get_or_404(user_id)
    if user.username == 'admin':
        flash('默认超级管理员不可删除', 'error')
        return redirect(url_for('admin.manage_users'))
    deleted_username = user.username
    db.session.delete(user)
    db.session.commit()
    current_app.logger.info('user_deleted by=%s target=%s', current_user.username, deleted_username)
    flash('鐢ㄦ埛宸插垹闄わ紒', 'success')
    return redirect(url_for('admin.manage_users'))

# ------------------- 鏍囨湰澶х被绠＄悊 ------------------- #
# 鍒楄〃灞曠ず
@admin_bp.route('/categories')
@login_required
@admin_required
def list_categories():
    categories = SpecimenCategory.query.order_by(SpecimenCategory.id.desc()).all()
    return render_template('admin/admin_category_management.html', categories=categories)

# 鏂板澶х被
@admin_bp.route('/categories/add', methods=['POST'])
@login_required
@admin_required
def add_category():
    name = request.form.get('name')
    code = request.form.get('code')
    description = request.form.get('description')
    if not name:
        flash('绫诲埆鍚嶇О涓嶈兘涓虹┖', 'warning')
        return redirect(url_for('admin.list_categories'))
    if not code:
        flash('唯一性代码不能为空', 'warning')
        return redirect(url_for('admin.list_categories'))

    if SpecimenCategory.query.filter_by(name=name).first():
        flash('璇ョ被鍒凡瀛樺湪', 'warning')
        return redirect(url_for('admin.list_categories'))
    if SpecimenCategory.query.filter_by(code=code).first():
        flash('璇ュ敮涓€鎬т唬鐮佸凡瀛樺湪', 'warning')
        return redirect(url_for('admin.list_categories'))

    category = SpecimenCategory(
        name=name,
        code=code,
        description=description,
        created_by=current_user.username,
        updated_by=current_user.username,
        created_at=datetime.now(),
        updated_at=datetime.now()
    )

    # 澶勭悊灏侀潰鍥剧墖涓婁紶
    image_file = request.files.get('image')
    if image_file and FileHandler.check_file(image_file):
        try:
            ext = image_file.filename.rsplit('.', 1)[1].lower()
            filename = f"{name}.{ext}"
            relative_path = FileHandler.save_file(image_file, folder_key='UPLOAD_FOLDER_CATEGORY_IMAGES', filename=filename)
            category.image = relative_path
        except IOError as e:
            flash(f'灏侀潰鍥剧墖淇濆瓨澶辫触: {e}', 'error')
            return redirect(url_for('admin.list_categories'))

    db.session.add(category)
    db.session.commit()
    current_app.logger.info('category_created by=%s name=%s code=%s', current_user.username, category.name, category.code)
    flash('澶х被娣诲姞鎴愬姛', 'success')
    return redirect(url_for('admin.list_categories'))

# 缂栬緫澶х被
# 缂栬緫澶х被锛堜粎POST锛?
@admin_bp.route('/categories/edit/<int:id>', methods=['POST'])
@login_required
@admin_required
def edit_category(id):
    category = SpecimenCategory.query.get_or_404(id)
    new_name = request.form.get('name')
    new_code = request.form.get('code')
    description = request.form.get('description')
    if not new_name:
        flash('绫诲埆鍚嶇О涓嶈兘涓虹┖', 'warning')
        return redirect(url_for('admin.list_categories'))
    if not new_code:
        flash('唯一性代码不能为空', 'warning')
        return redirect(url_for('admin.list_categories'))

    # 妫€鏌ユ槸鍚﹂噸澶?
    exists = SpecimenCategory.query.filter(
        SpecimenCategory.name == new_name,
        SpecimenCategory.id != id
    ).first()
    if exists:
        flash('璇ョ被鍒悕绉板凡瀛樺湪', 'warning')
        return redirect(url_for('admin.list_categories'))
    code_exists = SpecimenCategory.query.filter(
        SpecimenCategory.code == new_code,
        SpecimenCategory.id != id
    ).first()
    if code_exists:
        flash('璇ュ敮涓€鎬т唬鐮佸凡瀛樺湪', 'warning')
        return redirect(url_for('admin.list_categories'))

    category.name = new_name
    category.code = new_code
    category.description = description
    category.updated_by = current_user.username
    category.updated_at = datetime.now()

    # 澶勭悊灏侀潰鍥剧墖涓婁紶锛堟浛鎹㈡棫鍥撅級
    image_file = request.files.get('image')
    if image_file and FileHandler.check_file(image_file):
        try:
            # 鍒犻櫎鏃у浘鐗?
            if category.image:
                old_path = os.path.join(current_app.config['UPLOAD_FOLDER_CATEGORY_IMAGES'], os.path.basename(category.image))
                if os.path.exists(old_path):
                    os.remove(old_path)
            ext = image_file.filename.rsplit('.', 1)[1].lower()
            filename = f"{new_name}.{ext}"
            relative_path = FileHandler.save_file(image_file, folder_key='UPLOAD_FOLDER_CATEGORY_IMAGES', filename=filename)
            category.image = relative_path
        except IOError as e:
            flash(f'灏侀潰鍥剧墖淇濆瓨澶辫触: {e}', 'error')
            return redirect(url_for('admin.list_categories'))

    db.session.commit()
    current_app.logger.info('category_updated by=%s id=%s name=%s code=%s', current_user.username, category.id, category.name, category.code)
    flash('澶х被淇敼鎴愬姛', 'success')
    return redirect(url_for('admin.list_categories'))

# 鍒犻櫎澶х被
@admin_bp.route('/categories/delete/<int:id>', methods=['POST'])
@login_required
@admin_required
def delete_category(id):
    category = SpecimenCategory.query.get_or_404(id)
    if category.specimens:  # 濡傛灉鏈夊叧鑱旀爣鏈?
        flash('该分类下仍有关联标本，无法删除', 'danger')
        return redirect(url_for('admin.list_categories'))
    # 鍒犻櫎灏侀潰鍥剧墖鏂囦欢
    img_path = None
    if category.image:
        img_path = os.path.join(current_app.config['UPLOAD_FOLDER_CATEGORY_IMAGES'], os.path.basename(category.image))
    category_name = category.name
    db.session.delete(category)
    db.session.commit()
    if img_path and os.path.exists(img_path):
        os.remove(img_path)
    current_app.logger.info('category_deleted by=%s id=%s name=%s', current_user.username, id, category_name)
    flash('澶х被鍒犻櫎鎴愬姛', 'success')
    return redirect(url_for('admin.list_categories'))

# ==================== 鏍囨湰淇℃伅缁存姢 ==================== #
@admin_bp.route('/specimens')
@login_required
@admin_required
def manage_specimens():
    category_id = request.args.get('category_id', type=int)
    q = request.args.get('q', '').strip()
    categories = SpecimenCategory.query.all()

    query = Specimen.query.options(selectinload(Specimen.images))
    if category_id:
        query = query.filter_by(category_id=category_id)
    if q:
        search_filter = build_specimen_search_filter(q)
        if search_filter is not None:
            query = query.filter(search_filter)

    page = request.args.get('page', 1, type=int)
    per_page = 30
    pagination = query.order_by(Specimen.id.desc()) \
        .paginate(page=page, per_page=per_page, error_out=False)
    specimens = pagination.items

    # 鏋勫缓鏍囨湰 JSON 鏁版嵁锛屼緵缂栬緫妯℃€佹浣跨敤
    def fmt_dt(dt_val):
        if not dt_val:
            return ''
        from app.utils.time_utils import CHINA_TZ
        if dt_val.tzinfo is not None:
            dt_val = dt_val.astimezone(CHINA_TZ)
        else:
            dt_val = CHINA_TZ.localize(dt_val)
        return dt_val.strftime('%Y-%m-%dT%H:%M')

    specimens_json = {}
    for s in specimens:
        specimens_json[str(s.id)] = {
            'id': s.id,
            'category_id': s.category_id,
            'specimen_number': s.specimen_number,
            'chinese_name': s.chinese_name or '',
            'latin_name': s.latin_name or '',
            'alias': s.alias or '',
            'phylum': s.phylum or '',
            'class_name': s.class_name or '',
            'order_name': s.order_name or '',
            'family': s.family or '',
            'genus': s.genus or '',
            'species': s.species or '',
            'collector': s.collector or '',
            'collect_time': fmt_dt(s.collect_time),
            'collect_location': s.collect_location or '',
            'longitude': float(s.longitude) if s.longitude is not None else None,
            'latitude': float(s.latitude) if s.latitude is not None else None,
            'appraiser': s.appraiser or '',
            'appraisal_time': fmt_dt(s.appraisal_time),
            'other_info': s.other_info if isinstance(s.other_info, str) else (json.dumps(s.other_info, ensure_ascii=False) if s.other_info else None),
            'images': [{'id': img.id, 'url': url_for('static', filename=img.image_path)}
                       for img in s.images]
        }

    return render_template('admin/admin_specimen_management.html',
                           specimens=specimens,
                           specimens_json=specimens_json,
                           categories=categories,
                           current_category_id=category_id,
                           search_query=q,
                           pagination=pagination)

# ==================== 涓婁紶鏍囨湰 ===================== #
@admin_bp.route('/upload_specimen', methods=['GET', 'POST'])
@login_required
@admin_required
def upload_specimen():
    if request.method == 'POST':
        # ==================== 鑾峰彇琛ㄥ崟鏁版嵁 ==================== #
        category_id = request.form.get('category_id')  # 鏍囨湰澶х被 ID
        specimen_number = request.form.get('specimen_number')  # 鏍囨湰缂栧彿
        chinese_name = request.form.get('chinese_name') or None
        latin_name = request.form.get('latin_name')
        alias = request.form.get('alias') or None
        phylum = request.form.get('phylum') or None
        class_name = request.form.get('class_name') or None
        order_name = request.form.get('order') or None
        family = request.form.get('family') or None
        genus = request.form.get('genus') or None
        species = request.form.get('species') or None
        collector = request.form.get('collector') or None

        # ---- 蹇呭～瀛楁楠岃瘉锛堥棬/绾?鐩?绉?灞烇級 ----
        required_fields = [
            ('phylum', '门'),
            ('class_name', '纲'),
            ('order', '目'),
            ('family', '科'),
            ('genus', '属'),
        ]
        missing = []
        for field_name, label in required_fields:
            val = (request.form.get(field_name) or '').strip()
            if not val:
                missing.append(label)
        if missing:
            flash(f"以下必填字段不能为空：{'、'.join(missing)}", 'error')
            return redirect(url_for('admin.upload_specimen'))

        # ---- 鏍囨湰缂栧彿闈炵┖ + 鍞竴鎬ф鏌?----
        if not specimen_number or not specimen_number.strip():
            flash("鏍囨湰缂栧彿涓嶈兘涓虹┖", 'error')
            return redirect(url_for('admin.upload_specimen'))

        existing = Specimen.query.filter_by(specimen_number=specimen_number.strip()).first()
        if existing:
            flash(f"标本编号 '{specimen_number}' 已存在，请使用其他编号", 'error')
            return redirect(url_for('admin.upload_specimen'))

        # 閲囬泦鏃堕棿锛堝吋瀹?date 鍜?datetime-local 涓ょ鏍煎紡锛?
        collect_time = _parse_datetime_form_value(request.form.get('collect_time'), default=now())

        collect_location = request.form.get('collect_location') or None

        # ==================== 缁忕含搴﹀鐞?==================== #
        longitude = _parse_float_form_value(request.form.get('longitude'))
        latitude = _parse_float_form_value(request.form.get('latitude'))

        # ==================== 閴村畾淇℃伅 ==================== #
        appraiser = request.form.get('appraiser') or None
        appraisal_time = _parse_datetime_form_value(request.form.get('appraisal_time'))

        # ==================== 鍏朵粬淇℃伅锛圝SON 鎴栨枃鏈級 ==================== #
        other_info = request.form.get('other_info')
        try:
            other_info_json = json.loads(other_info) if other_info else None
        except Exception:
            other_info_json = other_info or None

        # ==================== 鍒涘缓 Specimen 瀹炰緥 ==================== #
        category_id_int = _parse_int_form_value(category_id)
        if category_id_int is None:
            flash("请选择有效的标本大类", 'error')
            return redirect(url_for('admin.upload_specimen'))
        specimen = Specimen(
            category_id=category_id_int,
            specimen_number=specimen_number,
            chinese_name=chinese_name,
            latin_name=latin_name,
            alias=alias,
            phylum=phylum,
            class_name=class_name,
            order_name=order_name,
            family=family,
            genus=genus,
            species=species,
            collector=collector,
            collect_time=collect_time,
            collect_location=collect_location,
            longitude=longitude,
            latitude=latitude,
            appraiser=appraiser,
            appraisal_time=appraisal_time,
            other_info=other_info_json,
            created_by=current_user.username,
            updated_by=current_user.username,
            # 鈿狅笍 涓嶅啀鎵嬪姩浼?created_at/updated_at锛屼娇鐢ㄦā鍨嬮粯璁?now()
        )
        db.session.add(specimen)
        db.session.flush()  # 鍏堣幏鍙?specimen.id锛屽皻鏈彁浜や簨鍔?

        # ==================== 澶勭悊澶氬紶鍥剧墖锛堜互鏍囨湰鍙峰懡鍚嶏級 ==================== #
        images = request.files.getlist('images')
        safe_number = re.sub(r'[\\/:*?"<>|]', '_', specimen_number)
        saved_files = []  # 璺熻釜宸蹭繚瀛樼殑鏂囦欢璺緞锛宑ommit 澶辫触鏃舵竻鐞?
        for index, image in enumerate(images):
            if FileHandler.check_file(image):
                try:
                    ext = image.filename.rsplit('.', 1)[1].lower()
                    if len(images) == 1:
                        custom_filename = f"{safe_number}.{ext}"
                    else:
                        custom_filename = f"{safe_number}_{index + 1}.{ext}"
                    relative_path = FileHandler.save_file(
                        image, folder_key='UPLOAD_FOLDER_SPECIMEN_IMAGES',
                        filename=custom_filename,
                        relative_base=current_app.static_folder
                    )
                    # 璁板綍缁濆璺緞锛岀敤浜?commit 澶辫触鏃舵竻鐞?
                    abs_path = os.path.join(current_app.static_folder, relative_path.replace('/', os.sep))
                    saved_files.append(abs_path)
                except IOError as e:
                    db.session.rollback()
                    # 娓呯悊鏈宸蹭繚瀛樼殑鏂囦欢
                    for f in saved_files:
                        try:
                            os.remove(f)
                        except OSError:
                            pass
                    flash(f"鍥剧墖淇濆瓨澶辫触: {e}", 'error')
                    return redirect(url_for('admin.upload_specimen'))
                img = SpecimenImage(
                    specimen_id=specimen.id,
                    image_path=relative_path,
                    sort_order=index + 1
                )
                db.session.add(img)

        try:
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            # 娓呯悊宸蹭繚瀛樼殑瀛ょ珛鏂囦欢
            for f in saved_files:
                try:
                    os.remove(f)
                except OSError:
                    pass
            flash(f"标本编号 '{specimen_number}' 已存在（并发冲突）", 'error')
            return redirect(url_for('admin.upload_specimen'))

        flash("标本上传成功", 'success')
        current_app.logger.info(
            'specimen_created by=%s number=%s category_id=%s image_count=%s',
            current_user.username,
            specimen_number,
            category_id_int,
            len(specimen.images),
        )
        return redirect(url_for('admin.upload_specimen'))

    # ==================== GET 璇锋眰鏄剧ず涓婁紶椤甸潰 ==================== #
    categories = SpecimenCategory.query.all()
    category_codes = {str(cat.id): cat.code for cat in categories if cat.code}
    return render_template('admin/upload_specimen.html', categories=categories, category_codes=category_codes)

# ==================== 鎵归噺涓婁紶鏍囨湰 ==================== #

@admin_bp.route('/batch_upload')
@login_required
@admin_required
def batch_upload():
    """Render the batch upload page."""
    return render_template('admin/batch_upload.html')


@admin_bp.route('/batch_upload/template')
@login_required
@admin_required
def batch_upload_template():
    """Generate the Excel template used for batch upload."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    # 鍒犻櫎榛樿 Sheet
    wb.remove(wb.active)

    categories = SpecimenCategory.query.order_by(SpecimenCategory.id).all()

    if not categories:
        flash('璇峰厛鍒涘缓鏍囨湰澶х被鍚庡啀涓嬭浇妯℃澘', 'error')
        return redirect(url_for('admin.batch_upload'))

    HEADERS = [
        '序号', '标本编号', '中文名', '拉丁名', '别名',
        '门', '纲', '目', '科', '属', '种',
        '采集人', '采集时间', '采集地点', '经度', '纬度',
        '鉴定人', '鉴定时间', '图片'
    ]

    # 琛ㄥご鏍峰紡
    header_font = Font(name='寰蒋闆呴粦', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='2C5F7C', end_color='2C5F7C', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    example_font = Font(name='寰蒋闆呴粦', size=10)
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0'),
    )

    for cat in categories:
        ws = wb.create_sheet(title=cat.name)

        # 鍐欏叆琛ㄥご锛堜粎琛ㄥご锛屼笉鍚ず渚嬫暟鎹锛?
        for col_idx, header in enumerate(HEADERS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 璁剧疆鍒楀
        col_widths = [6, 22, 18, 24, 14, 14, 14, 14, 14, 14, 14, 10, 14, 22, 14, 14, 10, 14, 22]
        for col_idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

        # Sheet 鏍囩棰滆壊锛堢粰姣忎釜澶х被涓€涓笉鍚岄鑹诧級
        tab_colors = ['FF2C5F7C', 'FF27AE60', 'FFB8934E', 'FF2D8A7B', 'FF8E44AD',
                      'FFE67E22', 'FF3498DB', 'FF1ABC9C', 'FFE74C3C', 'FFF39C12']
        color_idx = categories.index(cat) % len(tab_colors)
        ws.sheet_properties.tabColor = tab_colors[color_idx]

    # 杈撳嚭鍒?BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    wb.close()

    from flask import send_file
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='鏍囨湰鎵归噺瀵煎叆妯℃澘.xlsx'
    )


@admin_bp.route('/batch_upload/validate', methods=['POST'])
@login_required
@admin_required
def batch_upload_validate():
    """Validate uploaded ZIP package and return AJAX validation results."""
    # ---- 妫€鏌ユ枃浠?----
    zip_file = request.files.get('zip_file')
    if not zip_file or not zip_file.filename:
        return jsonify({'success': False, 'error': '璇烽€夋嫨 ZIP 鏂囦欢'})

    if not zip_file.filename.lower().endswith('.zip'):
        return jsonify({'success': False, 'error': '浠呮敮鎸?.zip 鏍煎紡鏂囦欢'})

    # 妫€鏌ユ枃浠跺ぇ灏?
    max_zip_size = current_app.config.get('MAX_ZIP_SIZE', 100 * 1024 * 1024)
    zip_file.seek(0, os.SEEK_END)
    file_size = zip_file.tell()
    zip_file.seek(0)
    if file_size > max_zip_size:
        max_mb = max_zip_size // (1024 * 1024)
        return jsonify({'success': False, 'error': f'鏂囦欢杩囧ぇ锛屾渶澶ф敮鎸?{max_mb}MB'})

    # ---- 鍒涘缓鍞竴涓存椂鐩綍 ----
    temp_dir = current_app.config.get('TEMP_UPLOAD_DIR')
    if not temp_dir:
        temp_dir = os.path.join(os.path.dirname(__file__), '..', 'temp_uploads')
    session_id = uuid.uuid4().hex
    extract_dir = os.path.join(temp_dir, session_id)
    os.makedirs(extract_dir, exist_ok=True)

    try:
        # ---- 瑙ｅ帇 ZIP ----
        with zipfile.ZipFile(zip_file) as zf:
            # 瀹夊叏妫€鏌ワ細闃叉 ZIP 鐐稿脊
            total_uncompressed = sum(info.file_size for info in zf.infolist())
            if total_uncompressed > 500 * 1024 * 1024:  # 500MB 涓婇檺
                shutil.rmtree(extract_dir, ignore_errors=True)
                return jsonify({'success': False, 'error': 'ZIP 解压后内容过大（超过 500MB）'})

            zf.extractall(extract_dir)

        # ---- 鏌ユ壘 Excel 鏂囦欢 ----
        xlsx_files = []
        for root, dirs, files in os.walk(extract_dir):
            for f in files:
                if f.lower().endswith('.xlsx') and not f.startswith('~$'):
                    xlsx_files.append(os.path.join(root, f))

        if not xlsx_files:
            shutil.rmtree(extract_dir, ignore_errors=True)
            return jsonify({'success': False, 'error': 'ZIP 涓湭鎵惧埌 .xlsx 鏂囦欢'})

        # 鍙栫涓€涓壘鍒扮殑 xlsx锛堝鏈夊涓紝浼樺厛鍙栨牴鐩綍涓嬬殑锛?
        xlsx_files.sort(key=lambda p: p.count(os.sep))
        xlsx_path = xlsx_files[0]

        # ---- 鍑嗗鍒嗙被鏄犲皠鍜屽凡鏈夌紪鍙?----
        all_categories = SpecimenCategory.query.all()
        categories_by_name = {cat.name: cat for cat in all_categories}

        existing_numbers = {s[0] for s in Specimen.query.with_entities(Specimen.specimen_number).all()}

        # ---- 鎵ц涓夊眰鏍￠獙 ----
        max_image_size = current_app.config.get('MAX_IMAGE_SIZE', 1 * 1024 * 1024)
        result = parse_workbook(
            file_path=xlsx_path,
            categories_by_name=categories_by_name,
            existing_numbers=existing_numbers,
            image_dir=extract_dir,  # ZIP 瑙ｅ帇鐩綍鍗冲浘鐗囨墍鍦ㄧ洰褰?
            max_image_size=max_image_size
        )

        # ---- 灏嗘牎楠岀粨鏋滃瓨鍏ヤ复鏃?JSON 鏂囦欢锛堥伩鍏?session cookie 婧㈠嚭锛?----
        valid_rows = [r for r in result['rows'] if r['is_valid']]
        cache_file = os.path.join(extract_dir, '_batch_cache.json')
        with open(cache_file, 'w', encoding='utf-8') as f:
            json.dump({'valid_rows': valid_rows, 'session_id': session_id}, f, ensure_ascii=False)

        # 鍙湪 session 涓瓨鍌?session_id锛堢敤浜庡悗缁鍏ユ椂瀹氫綅鏂囦欢锛?
        session['batch_session_id'] = session_id

        # ---- 鏋勫缓 JSON 鍝嶅簲 ----
        response_data = {
            'success': True,
            'session_id': session_id,
            'sheet_errors': result['sheet_errors'],
            'all_category_names': result.get('all_category_names', []),
            'sheet_details': result.get('sheet_details', {}),
            'rows': [],
            'summary': result['summary'],
        }

        for row in result['rows']:
            response_data['rows'].append({
                'sheet_name': row['sheet_name'],
                'row_index': row['row_index'],
                'specimen_number': row['data'].get('specimen_number') or '',
                'chinese_name': row['data'].get('chinese_name') or '',
                'latin_name': row['data'].get('latin_name') or '',
                'is_valid': row['is_valid'],
                'errors': row['errors'],
            })

        return jsonify(response_data)

    except zipfile.BadZipFile:
        shutil.rmtree(extract_dir, ignore_errors=True)
        return jsonify({'success': False, 'error': '无效的 ZIP 文件，无法解压'})
    except Exception as e:
        shutil.rmtree(extract_dir, ignore_errors=True)
        current_app.logger.error(f'鎵归噺涓婁紶鏍￠獙寮傚父: {e}', exc_info=True)
        return jsonify({'success': False, 'error': '服务器内部错误，请稍后再试'})


@admin_bp.route('/batch_upload/import', methods=['POST'])
@login_required
@admin_required
def batch_upload_import():
    """Import validated batch-upload rows and return AJAX import results."""
    session_id = session.get('batch_session_id')
    if not session_id:
        return jsonify({'success': False, 'error': '请先上传 ZIP 文件并完成校验'})

    # 浠庝复鏃?JSON 鏂囦欢涓鍙栨牎楠岄€氳繃鐨勬暟鎹?
    temp_dir = current_app.config.get('TEMP_UPLOAD_DIR')
    if not temp_dir:
        temp_dir = os.path.join(os.path.dirname(__file__), '..', 'temp_uploads')
    extract_dir = os.path.join(temp_dir, session_id)

    if not os.path.isdir(extract_dir):
        return jsonify({'success': False, 'error': '临时文件已过期，请重新上传 ZIP 文件'})

    cache_file = os.path.join(extract_dir, '_batch_cache.json')
    if not os.path.exists(cache_file):
        return jsonify({'success': False, 'error': '校验数据已过期，请重新上传 ZIP 文件'})

    with open(cache_file, 'r', encoding='utf-8') as f:
        cache_data = json.load(f)

    valid_rows = cache_data.get('valid_rows', [])
    if not valid_rows:
        return jsonify({'success': False, 'error': '没有可导入的数据，请重新上传 ZIP 文件并校验'})

    upload_folder = current_app.config.get('UPLOAD_FOLDER_SPECIMEN_IMAGES')

    success_count = 0
    failed_rows = []

    for row in valid_rows:
        data = row['data']
        specimen_number = (data.get('specimen_number') or '').strip()
        category_id = row['category_id']

        try:
            # ---- 鏋勫缓 Specimen 瀵硅薄 ----
            specimen = Specimen(
                category_id=category_id,
                specimen_number=specimen_number,
                latin_name=data.get('latin_name'),
                chinese_name=data.get('chinese_name'),
                alias=data.get('alias'),
                phylum=data.get('phylum'),
                class_name=data.get('class_name'),
                order_name=data.get('order_name'),
                family=data.get('family'),
                genus=data.get('genus'),
                species=data.get('species'),
                collector=data.get('collector'),
                collect_location=data.get('collect_location'),
                appraiser=data.get('appraiser'),
                created_by=current_user.username,
                updated_by=current_user.username,
            )

            # ---- 閲囬泦鏃堕棿 ----
            collect_time_str = data.get('collect_time')
            if collect_time_str:
                from app.utils.excel_util import _parse_date
                dt = _parse_date(collect_time_str)
                if dt:
                    specimen.collect_time = CHINA_TZ.localize(dt)
                else:
                    specimen.collect_time = now()
            else:
                specimen.collect_time = now()

            # ---- 閴村畾鏃堕棿 ----
            appraisal_time_str = data.get('appraisal_time')
            if appraisal_time_str:
                from app.utils.excel_util import _parse_date
                dt = _parse_date(appraisal_time_str)
                if dt:
                    specimen.appraisal_time = CHINA_TZ.localize(dt)

            # ---- 缁忕含搴?----
            lon_str = data.get('longitude')
            if lon_str:
                try:
                    specimen.longitude = float(lon_str)
                except (ValueError, TypeError):
                    pass

            lat_str = data.get('latitude')
            if lat_str:
                try:
                    specimen.latitude = float(lat_str)
                except (ValueError, TypeError):
                    pass

            db.session.add(specimen)
            copied_files = []  # 璺熻釜宸插鍒剁殑鏂囦欢锛宑ommit 澶辫触鏃舵竻鐞?
            db.session.flush()  # 鑾峰彇 specimen.id

            # ---- 澶勭悊鍥剧墖 ----
            image_filenames = data.get('image_filenames')
            if image_filenames:
                from app.utils.excel_util import _split_image_filenames
                filenames = _split_image_filenames(image_filenames)
                safe_number = re.sub(r'[\\/:*?"<>|]', '_', specimen_number)

                for img_index, fname in enumerate(filenames):
                    fname = fname.strip()
                    if not fname:
                        continue

                    src_path = _find_image_in_dir(fname, extract_dir)
                    if not src_path:
                        continue  # 鍥剧墖涓嶅瓨鍦紝璺宠繃锛堟牎楠岄樁娈靛凡鎶ラ敊锛?

                    ext = fname.rsplit('.', 1)[-1].lower()
                    if ext not in current_app.config.get('ALLOWED_EXTENSIONS', {'png', 'jpg', 'jpeg', 'gif'}):
                        continue

                    # 鍛藉悕瑙勫垯涓庡崟鏉′笂浼犱竴鑷?
                    total_images = len([f for f in filenames if f.strip()])
                    if total_images == 1:
                        dest_filename = f"{safe_number}.{ext}"
                    else:
                        dest_filename = f"{safe_number}_{img_index + 1}.{ext}"

                    dest_path = os.path.join(upload_folder, dest_filename)

                    # 濡傛灉鐩爣鏂囦欢宸插瓨鍦紝杩藉姞搴忓彿
                    if os.path.exists(dest_path):
                        base = dest_filename.rsplit('.', 1)[0]
                        counter = 1
                        while os.path.exists(os.path.join(upload_folder, f"{base}_{counter}.{ext}")):
                            counter += 1
                        dest_filename = f"{base}_{counter}.{ext}"
                        dest_path = os.path.join(upload_folder, dest_filename)

                    # 澶嶅埗鏂囦欢
                    shutil.copy2(src_path, dest_path)
                    copied_files.append(dest_path)

                    # 璁＄畻鐩稿璺緞
                    rel_path = os.path.relpath(dest_path, current_app.static_folder).replace('\\', '/')

                    img = SpecimenImage(
                        specimen_id=specimen.id,
                        image_path=rel_path,
                        sort_order=img_index + 1
                    )
                    db.session.add(img)

            # ---- 鎻愪氦璇ヨ ----
            db.session.commit()
            success_count += 1

        except IntegrityError:
            db.session.rollback()
            # 娓呯悊宸插鍒剁殑瀛ょ珛鏂囦欢
            for f in copied_files:
                try:
                    os.remove(f)
                except OSError:
                    pass
            failed_rows.append({
                'sheet_name': row['sheet_name'],
                'row_index': row['row_index'],
                'specimen_number': specimen_number,
                'reason': f'标本编号 "{specimen_number}" 已存在（并发冲突）'
            })
        except Exception as e:
            db.session.rollback()
            # 娓呯悊宸插鍒剁殑瀛ょ珛鏂囦欢
            for f in copied_files:
                try:
                    os.remove(f)
                except OSError:
                    pass
            current_app.logger.error(f'鎵归噺瀵煎叆琛屽け璐? {e}', exc_info=True)
            failed_rows.append({
                'sheet_name': row['sheet_name'],
                'row_index': row['row_index'],
                'specimen_number': specimen_number,
                'reason': str(e)
            })

    # ---- 娓呯悊涓存椂鐩綍 ----
    try:
        shutil.rmtree(extract_dir, ignore_errors=True)
    except Exception:
        pass

    # ---- 娓呴櫎 session ----
    session.pop('batch_session_id', None)

    current_app.logger.info(
        'batch_import_completed by=%s imported=%s failed=%s session_id=%s',
        current_user.username,
        success_count,
        len(failed_rows),
        session_id
    )

    return jsonify({
        'success': True,
        'imported': success_count,
        'failed': len(failed_rows),
        'failed_rows': failed_rows,
    })


def _find_image_in_dir(filename: str, search_dir: str) -> str | None:
    """Recursively find an image file under the extracted batch directory."""
    # 鍏堝湪鏍圭洰褰曟煡鎵?
    direct_path = os.path.join(search_dir, filename)
    if os.path.isfile(direct_path):
        return direct_path

    # 閫掑綊鏌ユ壘
    for root, dirs, files in os.walk(search_dir):
        for f in files:
            if f == filename:
                return os.path.join(root, f)

    return None

# ==================== 缂栬緫鏍囨湰 ==================== #
@admin_bp.route('/specimens/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_specimen(id):
    specimen = Specimen.query.get_or_404(id)
    categories = SpecimenCategory.query.all()
    category_codes = {str(cat.id): cat.code for cat in categories if cat.code}

    if request.method == 'POST':
        old_specimen_number = specimen.specimen_number  # 淇敼鍓嶇殑缂栧彿蹇収
        category_id_int = _parse_int_form_value(request.form.get('category_id'))
        if category_id_int is None:
            flash("请选择有效的标本大类", 'error')
            return redirect(url_for('admin.edit_specimen', id=specimen.id))
        specimen.category_id = category_id_int
        specimen.specimen_number = request.form.get('specimen_number')
        specimen.chinese_name = request.form.get('chinese_name') or None
        specimen.latin_name = request.form.get('latin_name')
        specimen.alias = request.form.get('alias') or None
        specimen.phylum = request.form.get('phylum') or None
        specimen.class_name = request.form.get('class_name') or None
        specimen.order_name = request.form.get('order') or None
        specimen.family = request.form.get('family') or None
        specimen.genus = request.form.get('genus') or None
        specimen.species = request.form.get('species') or None
        specimen.collector = request.form.get('collector') or None

        # ---- 蹇呭～瀛楁楠岃瘉锛堥棬/绾?鐩?绉?灞烇級 ----
        required_fields = [
            ('phylum', '门'),
            ('class_name', '纲'),
            ('order', '目'),
            ('family', '科'),
            ('genus', '属'),
        ]
        missing = []
        for field_name, label in required_fields:
            val = (request.form.get(field_name) or '').strip()
            if not val:
                missing.append(label)
        if missing:
            flash(f"以下必填字段不能为空：{'、'.join(missing)}", 'error')
            return redirect(url_for('admin.edit_specimen', id=specimen.id))

        # ---- 鏍囨湰缂栧彿闈炵┖ + 鍞竴鎬ф鏌ワ紙鎺掗櫎鑷韩锛?----
        new_number = request.form.get('specimen_number', '').strip()
        if not new_number:
            flash("鏍囨湰缂栧彿涓嶈兘涓虹┖", 'error')
            return redirect(url_for('admin.edit_specimen', id=specimen.id))

        existing = Specimen.query.filter(
            Specimen.specimen_number == new_number,
            Specimen.id != specimen.id
        ).first()
        if existing:
            flash(f"标本编号 '{new_number}' 已存在，请使用其他编号", 'error')
            return redirect(url_for('admin.edit_specimen', id=specimen.id))

        collect_time = _parse_datetime_form_value(request.form.get('collect_time'))
        if collect_time is not None:
            specimen.collect_time = collect_time

        specimen.collect_location = request.form.get('collect_location') or None

        specimen.longitude = _parse_float_form_value(request.form.get('longitude'))
        specimen.latitude = _parse_float_form_value(request.form.get('latitude'))

        specimen.appraiser = request.form.get('appraiser') or None
        specimen.appraisal_time = _parse_datetime_form_value(request.form.get('appraisal_time'))

        other_info = request.form.get('other_info')
        try:
            specimen.other_info = json.loads(other_info) if other_info else None
        except Exception:
            specimen.other_info = other_info or None

        # ---- 濡傛灉鏍囨湰缂栧彿鍙樻洿锛岃绠楅噸鍛藉悕璁″垝锛堟殏涓嶆墽琛屾枃浠舵搷浣滐級 ----
        pending_renames = []   # [(old_path, new_path)]
        if old_specimen_number != specimen.specimen_number:
            current_app.logger.info(
                f"鏍囨湰缂栧彿鍙樻洿: '{old_specimen_number}' -> '{specimen.specimen_number}', 璁″垝閲嶅懡鍚嶅浘鐗?.."
            )
            upload_folder = current_app.config.get('UPLOAD_FOLDER_SPECIMEN_IMAGES')
            safe_new = re.sub(r'[\\/:*?"<>|]', '_', specimen.specimen_number)
            current_app.logger.info(f"鍥剧墖涓婁紶鐩綍: {upload_folder}, safe_new='{safe_new}'")

            # 鏋勫缓寰呭垹闄ゅ浘鐗?ID 闆嗗悎锛堣繖浜涗笉闇€瑕侀噸鍛藉悕锛岀◢鍚庝細鍒犳帀锛?
            delete_ids_raw = request.form.get('delete_image_ids', '')
            delete_id_set = set()
            if delete_ids_raw:
                delete_id_set = {int(x) for x in delete_ids_raw.split(',') if x.strip()}

            # 鍓╀綑鍥剧墖锛堜笉鍚湰娆¤鍒犻櫎鐨勶級
            remaining_images = [img for img in specimen.images if img.id not in delete_id_set]
            total_remaining = len(remaining_images)
            current_app.logger.info(
                f"鏍囨湰鍏?{len(specimen.images)} 寮犲浘鐗囷紝"
                f"待删除 {len(delete_id_set)} 张，剩余 {total_remaining} 张"
            )

            for img in remaining_images:
                old_basename = os.path.basename(img.image_path)
                old_full_path = os.path.join(upload_folder, old_basename)
                ext = old_basename.rsplit('.', 1)[-1].lower()

                # 鏍规嵁鍓╀綑鍥剧墖鏁伴噺鍐冲畾鍛藉悕鏍煎紡
                if total_remaining == 1:
                    new_basename = f"{safe_new}.{ext}"
                else:
                    new_basename = f"{safe_new}_{img.sort_order}.{ext}"

                new_full_path = os.path.join(upload_folder, new_basename)
                current_app.logger.info(
                    f"  鍥剧墖 id={img.id}: '{old_basename}' -> '{new_basename}' "
                    f"(total_remaining={total_remaining}, sort_order={img.sort_order})"
                )

                # 濡傛灉鏂版棫鏂囦欢鍚嶇浉鍚岋紝璺宠繃
                if old_basename == new_basename:
                    current_app.logger.info("文件名未变化，跳过重命名")
                    continue

                # 鍙褰曢噸鍛藉悕璁″垝锛屾殏涓嶆墽琛岋紙绛?DB commit 鎴愬姛鍚庡啀鎿嶄綔鏂囦欢锛?
                if os.path.exists(old_full_path):
                    if os.path.exists(new_full_path):
                        current_app.logger.warning(f"鐩爣鏂囦欢宸插瓨鍦紝璺宠繃閲嶅懡鍚? {new_full_path}")
                        continue
                    pending_renames.append((old_full_path, new_full_path))
                    current_app.logger.info(f"  宸插姞鍏ラ噸鍛藉悕闃熷垪: {old_full_path} -> {new_full_path}")
                else:
                    current_app.logger.warning(f"  婧愭枃浠朵笉瀛樺湪锛屼粎鏇存柊鏁版嵁搴撹矾寰? {old_full_path}")

                # 鏇存柊鏁版嵁搴撲腑鐨勮矾寰勶紙鍏堟敼 DB锛屾枃浠舵搷浣滃湪 commit 涔嬪悗锛?
                rel_dir = os.path.dirname(img.image_path)
                img.image_path = os.path.join(rel_dir, new_basename).replace('\\', '/')
                current_app.logger.info(f"  鏁版嵁搴撹矾寰勫凡鏇存柊: '{img.image_path}'")

        # 澶勭悊鍥剧墖鍒犻櫎 鈥?鍙垹 DB 璁板綍锛屾枃浠舵搷浣滃欢杩熷埌 commit 涔嬪悗
        pending_deletions = []  # [file_path]
        delete_image_ids_str = request.form.get('delete_image_ids', '')
        if delete_image_ids_str:
            delete_image_ids = [int(x) for x in delete_image_ids_str.split(',') if x.strip()]
            for img_id in delete_image_ids:
                img = SpecimenImage.query.get(img_id)
                if img and img.specimen_id == specimen.id:
                    upload_folder = current_app.config.get('UPLOAD_FOLDER_SPECIMEN_IMAGES')
                    file_full_path = os.path.join(upload_folder, os.path.basename(img.image_path))
                    if os.path.exists(file_full_path):
                        pending_deletions.append(file_full_path)
                    db.session.delete(img)

        # 澶勭悊鏂板浘鐗囦笂浼狅紙浠ユ爣鏈彿鍛藉悕锛?
        existing_count = SpecimenImage.query.filter_by(specimen_id=specimen.id).count()
        images = request.files.getlist('images')
        safe_number = re.sub(r'[\\/:*?"<>|]', '_', specimen.specimen_number)
        pending_new_files = []  # 鐢ㄤ簬 commit 澶辫触鏃舵竻鐞?
        for index, image in enumerate(images):
            if FileHandler.check_file(image):
                try:
                    ext = image.filename.rsplit('.', 1)[1].lower()
                    seq = existing_count + index + 1
                    if len(images) == 1 and existing_count == 0:
                        custom_filename = f"{safe_number}.{ext}"
                    else:
                        custom_filename = f"{safe_number}_{seq}.{ext}"
                    relative_path = FileHandler.save_file(
                        image, folder_key='UPLOAD_FOLDER_SPECIMEN_IMAGES',
                        filename=custom_filename,
                        relative_base=current_app.static_folder
                    )
                except IOError as e:
                    db.session.rollback()
                    # 娓呯悊鏈宸蹭繚瀛樼殑鏂版枃浠?
                    for f in pending_new_files:
                        try:
                            os.remove(f)
                        except OSError:
                            pass
                    flash(f"鍥剧墖淇濆瓨澶辫触: {e}", 'error')
                    return redirect(url_for('admin.edit_specimen', id=specimen.id))
                # 璁板綍缁濆璺緞锛岀敤浜?commit 澶辫触鏃舵竻鐞?
                abs_path = os.path.join(current_app.static_folder, relative_path.replace('/', os.sep))
                pending_new_files.append(abs_path)
                img = SpecimenImage(
                    specimen_id=specimen.id,
                    image_path=relative_path,
                    sort_order=existing_count + index + 1
                )
                db.session.add(img)

        specimen.updated_by = current_user.username
        specimen.updated_at = now()

        try:
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            # 娓呯悊鏈宸蹭繚瀛樼殑鏂版枃浠?
            for f in pending_new_files:
                try:
                    os.remove(f)
                except OSError:
                    pass
            flash(f"标本编号 '{specimen.specimen_number}' 已存在（并发冲突）", 'error')
            return redirect(url_for('admin.edit_specimen', id=specimen.id))

        # ====== DB 鎻愪氦鎴愬姛锛屾墽琛屾枃浠舵搷浣?====== #
        # 1. 鎵ц寤舵湡閲嶅懡鍚?
        for old_path, new_path in pending_renames:
            try:
                os.rename(old_path, new_path)
                current_app.logger.info(f"閲嶅懡鍚嶆垚鍔? {old_path} -> {new_path}")
            except OSError as e:
                current_app.logger.warning(f"閲嶅懡鍚嶅け璐ワ紙鏁版嵁搴撳凡鏇存柊锛? {old_path} -> {new_path}, 閿欒: {e}")

        # 2. 鎵ц寤舵湡鍒犻櫎
        for file_path in pending_deletions:
            try:
                os.remove(file_path)
                current_app.logger.info(f"鍒犻櫎鏂囦欢鎴愬姛: {file_path}")
            except OSError as e:
                current_app.logger.warning(f"鍒犻櫎鏂囦欢澶辫触: {file_path}, 閿欒: {e}")

        current_app.logger.info('specimen_updated by=%s id=%s number=%s', current_user.username, specimen.id, specimen.specimen_number)
        flash('标本编辑成功', 'success')
        return redirect(url_for('admin.manage_specimens'))

    return render_template('admin/upload_specimen.html',
                           categories=categories,
                           specimen=specimen,
                           category_codes=category_codes)

# ==================== 鍒犻櫎鏍囨湰 ==================== #
@admin_bp.route('/specimens/delete/<int:id>', methods=['POST'])
@login_required
@admin_required
def delete_specimen(id):
    specimen = Specimen.query.get_or_404(id)
    specimen_number = specimen.specimen_number

    upload_folder = current_app.config.get('UPLOAD_FOLDER_SPECIMEN_IMAGES')
    for img in specimen.images:
        file_full_path = os.path.join(upload_folder, os.path.basename(img.image_path))
        if os.path.exists(file_full_path):
            try:
                os.remove(file_full_path)
            except OSError as e:
                current_app.logger.warning(f"鍒犻櫎鏍囨湰 {id} 鐨勫浘鐗囨枃浠跺け璐? {file_full_path}, 閿欒: {e}")

    db.session.delete(specimen)
    db.session.commit()

    current_app.logger.info('specimen_deleted by=%s id=%s number=%s', current_user.username, id, specimen_number)
    flash('鏍囨湰宸插垹闄わ紒', 'success')
    return redirect(url_for('admin.manage_specimens'))

