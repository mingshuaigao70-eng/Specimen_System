"""
Excel 批量上传解析与校验工具

支持多 Sheet 结构：
  - 每个 Sheet 名 = 标本大类名称（如 "浮游植物"）
  - 每个 Sheet 内的数据行按列名映射到 Specimen 模型字段

校验分为三层：
  第1层：Sheet 名是否在数据库大类中存在
  第2层：每行标本编号前缀是否匹配该 Sheet 对应的大类代码
  第3层：必填字段、日期格式、坐标有效性、图片文件存在性
"""
import re
import os
from datetime import datetime
from collections import OrderedDict

# Excel 列名 → 模型字段名
# 值为 None 的列仅用于人类阅读，解析时忽略
EXCEL_COLUMN_MAP = OrderedDict([
    ('序号',          None),
    ('标本编号',      'specimen_number'),
    ('中文名',        'chinese_name'),
    ('拉丁名',        'latin_name'),
    ('别名',          'alias'),
    ('门',            'phylum'),
    ('纲',            'class_name'),
    ('目',            'order_name'),
    ('科',            'family'),
    ('属',            'genus'),
    ('种',            'species'),
    ('采集人',        'collector'),
    ('采集时间',      'collect_time'),
    ('采集地点',      'collect_location'),
    ('经度',          'longitude'),
    ('纬度',          'latitude'),
    ('鉴定人',        'appraiser'),
    ('鉴定时间',      'appraisal_time'),
    ('图片',          'image_filenames'),
])

# 必填列（用于第3层校验）
REQUIRED_FIELDS = [
    'specimen_number', 'latin_name',
    'phylum', 'class_name', 'order_name', 'family', 'genus',
    'collect_time'
]

# 必需的 Excel 列名（对应上述必填字段，在表头校验时使用）
REQUIRED_COLUMN_NAMES = [
    '标本编号', '拉丁名',
    '门', '纲', '目', '科', '属',
    '采集时间'
]

# 允许的图片扩展名（与 config 中一致）
ALLOWED_IMAGE_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}


def extract_code_from_number(specimen_number: str) -> str | None:
    """
    从标本编号中提取大类代码。
    例: "HK-A-2026-99" → "A"
         "HK-BC-2025-001" → "BC"
    """
    if not specimen_number:
        return None
    m = re.match(r'HK-([A-Za-z]+)-', specimen_number.strip())
    return m.group(1).upper() if m else None


def _parse_date(value_str: str) -> datetime | None:
    """
    尝试将字符串解析为 datetime。
    支持格式：YYYY-MM-DD、YYYY/MM/DD、YYYY-M-D、YYYY/M/D（可带 HH:MM 时间部分）
    """
    if not value_str or not value_str.strip():
        return None
    value_str = value_str.strip()

    # 统一将斜杠和点号替换为横杠
    normalized = value_str.replace('/', '-').replace('.', '-')

    # 分离日期部分和时间部分
    time_part = ''
    if ' ' in normalized:
        parts = normalized.split(' ', 1)
        normalized = parts[0]
        time_part = ' ' + parts[1]

    # 将 YYYY-M-D 补零为 YYYY-MM-DD（确保跨平台兼容）
    try:
        date_segments = normalized.split('-')
        if len(date_segments) == 3:
            year, month, day = date_segments
            normalized = f"{int(year):04d}-{int(month):02d}-{int(day):02d}"
    except (ValueError, TypeError):
        pass

    value_to_parse = normalized + time_part
    for fmt in ('%Y-%m-%d %H:%M', '%Y-%m-%d'):
        try:
            return datetime.strptime(value_to_parse, fmt)
        except ValueError:
            continue
    return None


def _parse_float(value_str: str) -> tuple[float | None, bool]:
    """
    尝试将字符串解析为 float。
    返回 (value, is_valid) — is_valid 为 True 表示空白或有效数字。
    """
    if not value_str or not value_str.strip():
        return None, True
    try:
        return float(value_str.strip()), True
    except (ValueError, TypeError):
        return None, False


def validate_sheet_name(sheet_name: str, categories_by_name: dict) -> tuple[bool, str, object | None]:
    """
    第1层校验：Sheet 名是否匹配数据库中的大类名称。

    参数:
        sheet_name: 工作表名称（用户填写的）
        categories_by_name: {name: SpecimenCategory} 字典

    返回: (通过, 错误信息, category对象或None)
    """
    name = (sheet_name or '').strip()
    if not name:
        return False, '工作表名称为空', None

    category = categories_by_name.get(name)
    if category is None:
        valid_names = '、'.join(list(categories_by_name.keys()))
        return False, f'工作表「{name}」在标本大类中不存在。有效大类：{valid_names}', None

    return True, '', category


def validate_number_prefix(specimen_number: str, category: object) -> tuple[bool, str]:
    """
    第2层校验：编号前缀是否匹配大类代码。

    参数:
        specimen_number: 标本编号，如 "HK-A-2026-99"
        category: SpecimenCategory 对象（需要有 .code 属性）

    返回: (通过, 错误信息)
    """
    if not specimen_number or not specimen_number.strip():
        return False, '标本编号为空'

    code_in_number = extract_code_from_number(specimen_number.strip())
    expected_code = (category.code or '').strip().upper()

    if code_in_number is None:
        return False, f'标本编号格式不正确，应为 HK-{expected_code}-年份-序号（如 HK-{expected_code}-2026-001）'

    if code_in_number != expected_code:
        return False, (
            f'标本编号前缀 HK-{code_in_number} 与当前工作表对应大类'
            f'「{category.name}」(代码{expected_code}) 不匹配'
        )

    return True, ''


def validate_row(row_data: dict, image_dir: str | None, max_image_size: int = 1 * 1024 * 1024) -> list[str]:
    """
    第3层校验：字段内容级别。

    参数:
        row_data: {field_name: value} 字典
        image_dir: 图片文件所在目录（ZIP 解压目录），None 表示不检查图片存在性
        max_image_size: 单张图片最大大小（字节），默认 1MB

    返回: 错误信息列表，空列表表示通过
    """
    errors = []

    # ---- 必填字段 ----
    for field in REQUIRED_FIELDS:
        value = row_data.get(field)
        if value is None or (isinstance(value, str) and not value.strip()):
            # 找到对应的中文列名
            cn_name = _field_to_column_name(field)
            errors.append(f'缺少必填字段：{cn_name}')
            break  # 只报告第一个缺失字段，避免冗余

    # ---- 标本编号非空（已在必填中） ----
    # ---- 日期格式 ----
    for date_field in ('collect_time', 'appraisal_time'):
        value = row_data.get(date_field)
        if value and isinstance(value, str) and value.strip():
            if _parse_date(value) is None:
                cn_name = _field_to_column_name(date_field)
                errors.append(f'{cn_name}格式不正确：「{value}」，支持格式如 2026-06-11、2026/6/11、2026.06.11')

    # ---- 经度/纬度 ----
    for coord_field in ('longitude', 'latitude'):
        value = row_data.get(coord_field)
        if value and isinstance(value, str) and value.strip():
            _, ok = _parse_float(value)
            if not ok:
                cn_name = _field_to_column_name(coord_field)
                errors.append(f'{cn_name}不是有效数字')

    # ---- 图片文件校验 ----
    image_filenames = row_data.get('image_filenames')
    if image_filenames and image_dir:
        filenames = _split_image_filenames(image_filenames)
        for fname in filenames:
            fname = fname.strip()
            if not fname:
                continue
            # ① 检查扩展名
            ext = fname.rsplit('.', 1)[-1].lower() if '.' in fname else ''
            if ext not in ALLOWED_IMAGE_EXTENSIONS:
                errors.append(f'图片文件「{fname}」格式不支持（仅支持 jpg/png/gif）')
                continue
            # ② 检查文件是否存在
            full_path = os.path.join(image_dir, fname)
            if not os.path.isfile(full_path):
                errors.append(f'图片文件「{fname}」在 ZIP 包中未找到')
                continue
            # ③ 检查文件大小
            file_size = os.path.getsize(full_path)
            if file_size > max_image_size:
                max_mb = max_image_size / (1024 * 1024)
                actual_mb = file_size / (1024 * 1024)
                errors.append(f'图片文件「{fname}」过大（{actual_mb:.1f}MB，限制 {max_mb:.0f}MB）')
            # ④ 一次打开图片，完成损坏检测 + 内容启发式检查
            try:
                from PIL import Image
                img = Image.open(full_path)
                img.load()  # 实际加载图像数据，触发解码
                width, height = img.size
                # ④-a 尺寸检查：最小 200x200
                if width < 200 or height < 200:
                    errors.append(f'图片文件「{fname}」尺寸过小（{width}×{height}，最小 200×200）')
                # ④-b 宽高比检查：排除极端比例（如长截图）
                ratio = max(width, height) / max(min(width, height), 1)
                if ratio > 10:
                    errors.append(f'图片文件「{fname}」宽高比异常（{width}×{height}，比例 {ratio:.0f}:1）')
                img.close()
            except Exception:
                errors.append(f'图片文件「{fname}」已损坏或无法正确读取')
    elif not image_filenames or not (isinstance(image_filenames, str) and image_filenames.strip()):
        # ⑤ 图片列为空
        errors.append('缺少图片，请至少提供一张标本图片')

    return errors


def check_cross_sheet_duplicates(rows: list[dict]) -> list[str]:
    """
    跨 Sheet 检查标本编号在 Excel 内部是否重复。

    参数:
        rows: 所有已解析的行数据列表，每项含 sheet_name, row_index, data

    返回: 发现的重复错误列表（同时会修改对应行的 errors 字段）
    """
    seen = {}  # {specimen_number: (sheet_name, row_index)}
    duplicate_errors = []

    for row in rows:
        number = (row.get('data', {}).get('specimen_number') or '').strip()
        if not number:
            continue
        if number in seen:
            prev_sheet, prev_row = seen[number]
            msg = f'标本编号 {number} 与 Sheet「{prev_sheet}」第 {prev_row} 行重复'
            row.setdefault('errors', []).append(msg)
            duplicate_errors.append(msg)
        else:
            seen[number] = (row.get('sheet_name', ''), row.get('row_index', 0))

    return duplicate_errors


def check_db_duplicates(rows: list[dict], existing_numbers: set[str]) -> int:
    """
    检查标本编号是否在数据库中已存在。

    参数:
        rows: 所有已解析的行数据列表
        existing_numbers: 数据库中已存在的标本编号集合

    返回: 发现的冲突数量
    """
    count = 0
    for row in rows:
        number = (row.get('data', {}).get('specimen_number') or '').strip()
        if number and number in existing_numbers:
            row.setdefault('errors', []).append(f'标本编号 {number} 在数据库中已存在')
            count += 1
    return count


def parse_workbook(
    file_path: str,
    categories_by_name: dict,
    existing_numbers: set[str],
    image_dir: str | None = None,
    max_image_size: int = 1 * 1024 * 1024
) -> dict:
    """
    解析多 Sheet Excel 工作簿，完成全部三层校验。

    参数:
        file_path: .xlsx 文件的完整路径
        categories_by_name: {大类名称: SpecimenCategory} 字典
        existing_numbers: 数据库中已存在的所有标本编号集合
        image_dir: ZIP 解压后的图片目录（可选，用于校验图片文件存在性）
        max_image_size: 单张图片最大大小（字节），默认 1MB

    返回:
        {
            'sheet_errors': [{'sheet_name': str, 'error': str}],
            'rows': [
                {
                    'sheet_name': str,
                    'row_index': int,        # 该 Sheet 内的行号（1-based 数据行）
                    'category_id': int | None,
                    'category_code': str | None,
                    'data': {field_name: value},
                    'errors': [str],
                    'is_valid': bool
                }
            ],
            'summary': {
                'total_sheets': int,
                'valid_sheets': int,
                'total_rows': int,
                'valid_rows': int,
                'error_rows': int
            }
        }
    """
    from openpyxl import load_workbook

    wb = load_workbook(file_path, read_only=True, data_only=True)
    sheet_names = wb.sheetnames

    sheet_errors = []
    all_rows = []
    valid_category_sheets = {}  # {sheet_name: category_obj}

    # 先获取所有大类名称用于错误提示
    all_category_names = list(categories_by_name.keys())

    # ---- 第1层：遍历 Sheet，校验名称 ----
    for sheet_name in sheet_names:
        name = (sheet_name or '').strip()
        if not name:
            sheet_errors.append({'sheet_name': sheet_name or '(空)', 'error': '工作表名称为空，已跳过'})
            continue

        ok, err_msg, category = validate_sheet_name(name, categories_by_name)
        if not ok:
            sheet_errors.append({'sheet_name': name, 'error': err_msg})
            continue

        valid_category_sheets[name] = category

    # ---- 解析每个有效 Sheet 的行数据 ----
    # 全局列头映射（所有 Sheet 共用同一套列名）
    workbook_headers = None  # {col_index: field_name}

    for sheet_name in valid_category_sheets:
        ws = wb[sheet_name]
        category = valid_category_sheets[sheet_name]

        # 读取所有行（openpyxl read_only 模式下 ws 是迭代器）
        all_ws_rows = list(ws.iter_rows(values_only=True))
        if not all_ws_rows:
            continue

        # 第一行是表头
        header_row = all_ws_rows[0]
        if workbook_headers is None:
            # 首次遇到有效 Sheet，解析全局表头映射
            workbook_headers = _parse_headers(header_row)
            if not workbook_headers:
                sheet_errors.append({
                    'sheet_name': sheet_name,
                    'error': '表头行缺少任何有效列名，请确认 Excel 格式与模板一致'
                })
                continue

        # 解析数据行（第2行开始）
        for row_idx, row_values in enumerate(all_ws_rows[1:], start=2):
            row_data = _map_row_to_data(row_values, workbook_headers)
            errors = []

            # 跳过完全空行
            if not any(v for v in row_data.values() if v):
                continue

            # ---- 第2层：编号前缀校验 ----
            specimen_number = (row_data.get('specimen_number') or '').strip()
            if specimen_number:
                prefix_ok, prefix_err = validate_number_prefix(specimen_number, category)
                if not prefix_ok:
                    errors.append(prefix_err)
            # 如果编号为空，稍后第3层会报"缺少必填字段"

            # ---- 第3层：字段内容校验 ----
            content_errors = validate_row(row_data, image_dir, max_image_size)
            errors.extend(content_errors)

            all_rows.append({
                'sheet_name': sheet_name,
                'row_index': row_idx,
                'category_id': category.id,
                'category_code': category.code,
                'data': row_data,
                'errors': errors,
                'is_valid': len(errors) == 0,
            })

    wb.close()

    # ---- 跨 Sheet 重复检查 ----
    check_cross_sheet_duplicates(all_rows)

    # 跨 Sheet 重复检查后更新 is_valid
    for row in all_rows:
        row['is_valid'] = len(row['errors']) == 0

    # ---- 数据库重复检查 ----
    check_db_duplicates(all_rows, existing_numbers)

    # 数据库重复检查后再次更新 is_valid
    for row in all_rows:
        row['is_valid'] = len(row['errors']) == 0

    # ---- 汇总 ----
    total_sheets = len(sheet_names)
    valid_sheets = len(valid_category_sheets) - sum(
        1 for e in sheet_errors if e['sheet_name'] in valid_category_sheets
    )
    total_rows = len(all_rows)
    valid_rows = sum(1 for r in all_rows if r['is_valid'])
    error_rows = total_rows - valid_rows

    # ---- 每个有效工作表的统计 ----
    sheet_details = {}
    for sheet_name in valid_category_sheets:
        sheet_rows = [r for r in all_rows if r['sheet_name'] == sheet_name]
        sheet_valid = sum(1 for r in sheet_rows if r['is_valid'])
        sheet_error = len(sheet_rows) - sheet_valid
        sheet_details[sheet_name] = {
            'total_rows': len(sheet_rows),
            'valid_rows': sheet_valid,
            'error_rows': sheet_error,
        }

    return {
        'sheet_errors': sheet_errors,
        'rows': all_rows,
        'all_category_names': all_category_names,
        'sheet_details': sheet_details,
        'summary': {
            'total_sheets': total_sheets,
            'valid_sheets': valid_sheets,
            'total_rows': total_rows,
            'valid_rows': valid_rows,
            'error_rows': error_rows,
        }
    }


def _parse_headers(header_row: tuple) -> dict:
    """
    解析表头行，返回 {列索引: 模型字段名} 映射。
    表头应使用中文列名，通过 EXCEL_COLUMN_MAP 映射到模型字段。
    """
    col_map = {}
    for idx, cell_value in enumerate(header_row):
        if cell_value is None:
            continue
        col_name = str(cell_value).strip()
        field_name = EXCEL_COLUMN_MAP.get(col_name)
        if field_name:  # 仅映射有对应模型字段的列（None 值表示忽略列）
            col_map[idx] = field_name
    return col_map if col_map else None


def _map_row_to_data(row_values: tuple, col_map: dict) -> dict:
    """
    根据列映射将行数据转换为 {field_name: value} 字典。
    """
    data = {}
    for idx, field_name in col_map.items():
        value = row_values[idx] if idx < len(row_values) else None
        # 转为字符串（openpyxl 会返回各种类型）
        if value is not None:
            if isinstance(value, datetime):
                value = value.strftime('%Y-%m-%d %H:%M')
            else:
                value = str(value).strip()
                # 防止 CSV/Excel 公式注入：对以 = + - @ 开头的值添加单引号前缀
                if value and value[0] in '=+-@':
                    value = "'" + value
        data[field_name] = value if value else None
    return data


def _field_to_column_name(field_name: str) -> str:
    """将模型字段名反向映射为中文列名，用于错误提示。"""
    for cn_name, f_name in EXCEL_COLUMN_MAP.items():
        if f_name == field_name:
            return cn_name
    return field_name



def _split_image_filenames(value: str) -> list[str]:
    """
    拆分图片文件名字符串。
    支持逗号、分号、中文逗号/分号作为分隔符。
    """
    # 统一分隔符
    value = value.replace('；', ';').replace('，', ',')
    parts = []
    for part in value.split(','):
        part = part.strip()
        if part:
            parts.append(part)
    # 如果只有一个元素，尝试用分号拆分
    if len(parts) == 1:
        parts = [p.strip() for p in value.split(';') if p.strip()]
    return parts
